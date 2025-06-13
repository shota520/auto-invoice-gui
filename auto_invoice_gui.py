import tkinter as tk
from tkinter import filedialog, messagebox
import csv
from openpyxl import load_workbook
from datetime import datetime

template_map = {
    "請求書": "template_invoice.xlsx",
    "見積書": "template_estimate.xlsx",
    "納品書": "template_delivery.xlsx"
}

def generate_invoices(csv_path, template_name):
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        clients = {}
        for row in reader:
            name = row["name"]
            item = row["item"]
            quantity = int(row["quantity"])
            unit_price = int(row["unit_price"])
            if name not in clients:
                clients[name] = []
            clients[name].append((item, quantity, unit_price))

    today = datetime.today().strftime("%Y%m%d")
    template_file = template_map[template_name]

    for name, details in clients.items():
        wb = load_workbook(template_file)
        ws = wb.active

        ws["B3"] = f"{name} 株式会社"
        ws["B5"] = datetime.today().strftime("%Y-%m-%d")

        start_row = 7
        for i, (item, qty, price) in enumerate(details):
            row = start_row + i
            ws[f"A{row}"] = item
            ws[f"B{row}"] = qty
            ws[f"C{row}"] = price
            ws[f"D{row}"] = f"=B{row}*C{row}"

        filename = f"output_{template_name}_{name}_{today}.xlsx".replace("template_", "")
        wb.save(filename)

    messagebox.showinfo("完了", "帳票の出力が完了しました。")

def start_gui():
    def select_csv():
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if path:
            csv_path.set(path)

    def execute():
        generate_invoices(csv_path.get(), template_choice.get())

    root = tk.Tk()
    root.title("帳票出力カンタンツール")

    tk.Label(root, text="CSVファイルを選択:").grid(row=0, column=0, sticky="w")
    csv_path = tk.StringVar()
    tk.Entry(root, textvariable=csv_path, width=40).grid(row=0, column=1)
    tk.Button(root, text="参照", command=select_csv).grid(row=0, column=2)

    tk.Label(root, text="テンプレートを選択:").grid(row=1, column=0, sticky="w")
    template_choice = tk.StringVar(value="請求書")
    tk.OptionMenu(root, template_choice, *template_map.keys()).grid(row=1, column=1, sticky="w")

    tk.Button(root, text="出力実行", command=execute).grid(row=2, column=1, pady=10)

    root.mainloop()

if __name__ == "__main__":
    start_gui()
