from openpyxl import Workbook

def create_template(title, company_label, date_label, file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # ヘッダー
    ws["A1"] = title
    ws["A3"] = company_label
    ws["B3"] = "〇〇株式会社"
    ws["A5"] = date_label
    ws["B5"] = "2025-06-13"

    # 明細ヘッダー
    ws["A7"] = "商品名" if title == "請求書" else "項目" if title == "見積書" else "品目"
    ws["B7"] = "数量"
    ws["C7"] = "単価"
    ws["D7"] = "金額" if title == "請求書" else "小計" if title == "見積書" else "合計"

    # ダミー行（Excel関数を含める）
    ws["A8"] = "サービスA"
    ws["B8"] = 2
    ws["C8"] = 5000
    ws["D8"] = "=B8*C8"

    # 保存
    wb.save(file_name)
    print(f"{file_name} を作成しました。")

# 3種類作成
create_template("請求書", "会社名:", "日付:", "template_invoice.xlsx")
create_template("見積書", "御見積先:", "発行日:", "template_estimate.xlsx")
create_template("納品書", "納品先:", "納品日:", "template_delivery.xlsx")
